import os
import sys
import tempfile
import threading
import webbrowser

from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import pandas as pd
from scheduler import (
    generate_schedule,
    generate_praktikum,
    generate_from_db,
    detect_conflicts,
    ScheduleError,
)
from db import init_db, get_db


def resource_path(relative):
    """Resolve path to bundled resources (works in dev and PyInstaller)."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)


app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static"),
)
app.secret_key = "jadwal-farmasi-lokal"
init_db()

current_schedule = []
current_praktikum = []


def _parse_int(value, label, minimum=0):
    try:
        n = int(str(value).strip())
    except (ValueError, TypeError):
        raise ScheduleError(f"{label} harus berupa angka.")
    if n < minimum:
        raise ScheduleError(f"{label} minimal {minimum}.")
    return n


def _read_upload(file):
    """Baca file Excel/CSV jadi DataFrame, kolom diseragamkan huruf kecil."""
    name = (file.filename or "").lower()
    df = pd.read_csv(file) if name.endswith(".csv") else pd.read_excel(file)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


# ---------- Dashboard ----------
@app.route("/")
def dashboard():
    db = get_db()
    stats = {
        "angkatan": db.execute("SELECT COUNT(*) c FROM angkatan").fetchone()["c"],
        "ruang": db.execute("SELECT COUNT(*) c FROM ruang").fetchone()["c"],
        "dosen": db.execute("SELECT COUNT(*) c FROM dosen").fetchone()["c"],
        "matakuliah": db.execute("SELECT COUNT(*) c FROM matakuliah").fetchone()["c"],
    }
    db.close()
    return render_template("dashboard.html", stats=stats)


# ---------- Angkatan ----------
@app.route("/angkatan")
def angkatan_list():
    db = get_db()
    rows = db.execute("SELECT * FROM angkatan ORDER BY tahun DESC").fetchall()
    db.close()
    return render_template("angkatan.html", rows=rows)


@app.route("/angkatan/simpan", methods=["POST"])
def angkatan_simpan():
    jml_kelas = int(request.form.get("jml_kelas") or 12)
    jml_grup = (jml_kelas + 1) // 2  # grup praktikum = gabungan 2 kelas
    db = get_db()
    db.execute(
        "INSERT INTO angkatan (tahun, jml_kelas, jml_grup) VALUES (?,?,?)",
        (request.form["tahun"], jml_kelas, jml_grup),
    )
    db.commit(); db.close()
    flash("Angkatan ditambahkan.")
    return redirect(url_for("angkatan_list"))


@app.route("/angkatan/hapus/<int:id>")
def angkatan_hapus(id):
    db = get_db(); db.execute("DELETE FROM angkatan WHERE id=?", (id,)); db.commit(); db.close()
    return redirect(url_for("angkatan_list"))


@app.route("/angkatan/import", methods=["POST"])
def angkatan_import():
    try:
        df = _read_upload(request.files["file"])
        db = get_db()
        for _, r in df.iterrows():
            jml_kelas = int(r.get("jml_kelas", 12))
            db.execute("INSERT INTO angkatan (tahun, jml_kelas, jml_grup) VALUES (?,?,?)",
                       (str(r.get("tahun", "")), jml_kelas, (jml_kelas + 1) // 2))
        db.commit(); db.close()
        flash(f"{len(df)} angkatan diimpor.")
    except Exception as e:
        flash(f"Gagal impor: {e}")
    return redirect(url_for("angkatan_list"))


# ---------- Ruang ----------
@app.route("/ruang")
def ruang_list():
    db = get_db()
    rows = db.execute("SELECT * FROM ruang ORDER BY tipe, kode").fetchall()
    db.close()
    return render_template("ruang.html", rows=rows)


@app.route("/ruang/simpan", methods=["POST"])
def ruang_simpan():
    db = get_db()
    db.execute(
        "INSERT INTO ruang (kode, nama, kapasitas, tipe) VALUES (?,?,?,?)",
        (request.form["kode"], request.form.get("nama", ""), request.form.get("kapasitas", 0), request.form.get("tipe", "kuliah")),
    )
    db.commit(); db.close()
    flash("Ruang ditambahkan.")
    return redirect(url_for("ruang_list"))


@app.route("/ruang/hapus/<int:id>")
def ruang_hapus(id):
    db = get_db(); db.execute("DELETE FROM ruang WHERE id=?", (id,)); db.commit(); db.close()
    return redirect(url_for("ruang_list"))


@app.route("/ruang/import", methods=["POST"])
def ruang_import():
    try:
        df = _read_upload(request.files["file"])
        db = get_db()
        for _, r in df.iterrows():
            db.execute("INSERT INTO ruang (kode, nama, kapasitas, tipe) VALUES (?,?,?,?)",
                       (str(r.get("kode", "")), str(r.get("nama", "")), int(r.get("kapasitas", 0)), str(r.get("tipe", "kuliah"))))
        db.commit(); db.close()
        flash(f"{len(df)} ruang diimpor.")
    except Exception as e:
        flash(f"Gagal impor: {e}")
    return redirect(url_for("ruang_list"))


# ---------- Dosen ----------
@app.route("/dosen")
def dosen_list():
    db = get_db()
    rows = db.execute("SELECT * FROM dosen ORDER BY nama").fetchall()
    db.close()
    return render_template("dosen.html", rows=rows)


@app.route("/dosen/simpan", methods=["POST"])
def dosen_simpan():
    db = get_db()
    db.execute(
        "INSERT INTO dosen (nama, nidn, tidak_tersedia) VALUES (?,?,?)",
        (request.form["nama"], request.form.get("nidn", ""), request.form.get("tidak_tersedia", "")),
    )
    db.commit(); db.close()
    flash("Dosen ditambahkan.")
    return redirect(url_for("dosen_list"))


@app.route("/dosen/hapus/<int:id>")
def dosen_hapus(id):
    db = get_db(); db.execute("DELETE FROM dosen WHERE id=?", (id,)); db.commit(); db.close()
    return redirect(url_for("dosen_list"))


@app.route("/dosen/import", methods=["POST"])
def dosen_import():
    try:
        df = _read_upload(request.files["file"])
        db = get_db()
        for _, r in df.iterrows():
            db.execute("INSERT INTO dosen (nama, nidn, tidak_tersedia) VALUES (?,?,?)",
                       (str(r.get("nama", "")), str(r.get("nidn", "")), str(r.get("tidak_tersedia", ""))))
        db.commit(); db.close()
        flash(f"{len(df)} dosen diimpor.")
    except Exception as e:
        flash(f"Gagal impor: {e}")
    return redirect(url_for("dosen_list"))


# ---------- Mata Kuliah ----------
@app.route("/matakuliah")
def mk_list():
    db = get_db()
    rows = db.execute(
        "SELECT m.*, a.tahun, d.nama dosen FROM matakuliah m "
        "LEFT JOIN angkatan a ON m.angkatan_id=a.id LEFT JOIN dosen d ON m.dosen_id=d.id "
        "ORDER BY m.jenis, m.kode"
    ).fetchall()
    angkatan = db.execute("SELECT * FROM angkatan ORDER BY tahun DESC").fetchall()
    dosen = db.execute("SELECT * FROM dosen ORDER BY nama").fetchall()
    lab_names = [r["nama"] for r in db.execute(
        "SELECT DISTINCT nama FROM ruang WHERE tipe='lab' AND nama IS NOT NULL AND nama<>'' ORDER BY nama"
    )]
    db.close()
    return render_template("matakuliah.html", rows=rows, angkatan=angkatan, dosen=dosen, lab_names=lab_names)


@app.route("/matakuliah/simpan", methods=["POST"])
def mk_simpan():
    db = get_db()
    db.execute(
        "INSERT INTO matakuliah (kode, nama, sks, jenis, angkatan_id, dosen_id, lab) VALUES (?,?,?,?,?,?,?)",
        (request.form["kode"], request.form["nama"], request.form.get("sks", 2),
         request.form.get("jenis", "wajib"), request.form.get("angkatan_id") or None,
         request.form.get("dosen_id") or None, request.form.get("lab") or None),
    )
    db.commit(); db.close()
    flash("Mata kuliah ditambahkan.")
    return redirect(url_for("mk_list"))


@app.route("/matakuliah/hapus/<int:id>")
def mk_hapus(id):
    db = get_db(); db.execute("DELETE FROM matakuliah WHERE id=?", (id,)); db.commit(); db.close()
    return redirect(url_for("mk_list"))


@app.route("/matakuliah/import", methods=["POST"])
def mk_import():
    try:
        df = _read_upload(request.files["file"])
        db = get_db()
        ang = {str(a["tahun"]): a["id"] for a in db.execute("SELECT * FROM angkatan")}
        dos = {a["nama"]: a["id"] for a in db.execute("SELECT * FROM dosen")}
        for _, r in df.iterrows():
            db.execute("INSERT INTO matakuliah (kode, nama, sks, jenis, angkatan_id, dosen_id, lab) VALUES (?,?,?,?,?,?,?)",
                       (str(r.get("kode", "")), str(r.get("nama", "")), int(r.get("sks", 2)),
                        str(r.get("jenis", "wajib")), ang.get(str(r.get("angkatan", ""))), dos.get(str(r.get("dosen", ""))),
                        str(r.get("lab", "")) or None))
        db.commit(); db.close()
        flash(f"{len(df)} mata kuliah diimpor.")
    except Exception as e:
        flash(f"Gagal impor: {e}")
    return redirect(url_for("mk_list"))


# ---------- Generate dari Master Data ----------
@app.route("/jadwal", methods=["GET", "POST"])
def jadwal():
    global current_schedule, current_praktikum
    error = None
    if request.method == "POST":
        db = get_db()
        angkatan = [dict(r) for r in db.execute("SELECT * FROM angkatan")]
        rooms = [r["kode"] for r in db.execute("SELECT * FROM ruang WHERE tipe='kuliah' ORDER BY kode")]
        labs = [dict(r) for r in db.execute("SELECT kode, nama FROM ruang WHERE tipe='lab' ORDER BY kode")]
        mk = db.execute(
            "SELECT m.kode, m.jenis, m.angkatan_id, m.lab, d.nama dosen FROM matakuliah m "
            "LEFT JOIN dosen d ON m.dosen_id=d.id"
        ).fetchall()
        db.close()
        mk_wajib = [dict(r) for r in mk if r["jenis"] == "wajib"]
        mk_pilihan = [dict(r) for r in mk if r["jenis"] == "pilihan"]
        mk_praktikum = [dict(r) for r in mk if r["jenis"] == "praktikum"]
        try:
            current_schedule, current_praktikum = generate_from_db(
                angkatan=angkatan, rooms=rooms, labs=labs,
                mk_wajib=mk_wajib, mk_pilihan=mk_pilihan, mk_praktikum=mk_praktikum,
            )
        except ScheduleError as e:
            error = str(e)
    conflicts = detect_conflicts(current_schedule, current_praktikum)
    return render_template(
        "jadwal.html", schedule=current_schedule, praktikum=current_praktikum,
        error=error, conflicts=conflicts,
    )


# ---------- Simpan edit manual jadwal ----------
@app.route("/jadwal/simpan", methods=["POST"])
def jadwal_simpan():
    global current_schedule, current_praktikum
    for key, val in request.form.items():
        parts = key.split("|")
        if len(parts) != 3:
            continue
        typ, idx, col = parts
        try:
            idx = int(idx)
        except ValueError:
            continue
        target = current_schedule if typ == "k" else current_praktikum
        if 0 <= idx < len(target) and col in target[idx]:
            target[idx][col] = val.strip() or "-"
    flash("Perubahan disimpan.", "ok")
    return redirect(url_for("jadwal"))


# ---------- Generate manual (parameter) ----------
@app.route("/generate", methods=["GET", "POST"])
def index():
    global current_schedule, current_praktikum
    error = None
    prak_matrix = []

    if request.method == "POST":
        try:
            num_rooms = _parse_int(request.form.get("num_rooms"), "Jumlah ruangan", 1)
            num_pilihan = _parse_int(request.form.get("num_pilihan"), "Jumlah MK pilihan", 0)
            kelas_mkw = _parse_int(request.form.get("kelas_mkw"), "Kelas per MK wajib", 1)
            kelas_mkp = _parse_int(request.form.get("kelas_mkp"), "Kelas per MK pilihan", 1)

            raw = (request.form.get("distribusi") or "").split(",")
            distribusi = [_parse_int(x, "Distribusi MK wajib", 0) for x in raw if x.strip()]
            if not distribusi:
                raise ScheduleError("Distribusi MK wajib tidak boleh kosong.")

            raw_lab = (request.form.get("lab_kapasitas") or "").split(",")
            lab_kapasitas = [_parse_int(x, "Kapasitas lab", 1) for x in raw_lab if x.strip()]
            if not lab_kapasitas:
                raise ScheduleError("Kapasitas lab tidak boleh kosong.")

            # Matriks praktikum: satu input angka per sel (p_<angkatan>_<lab>).
            distribusi_prak = []
            for a in range(len(distribusi)):
                distribusi_prak.append([
                    _parse_int(request.form.get(f"p_{a}_{l}") or 0, "MK praktikum", 0)
                    for l in range(len(lab_kapasitas))
                ])
            prak_matrix = distribusi_prak

            # MKP untuk angkatan 2024 & 2025 = 2 angkatan terakhir
            angkatan_pilihan = [len(distribusi) - 2, len(distribusi) - 1]
            angkatan_pilihan = [a for a in angkatan_pilihan if a >= 0]

            current_schedule = generate_schedule(
                num_rooms=num_rooms,
                distribusi_semester=distribusi,
                num_mk_pilihan=num_pilihan,
                kelas_per_mkw=kelas_mkw,
                kelas_per_mkp=kelas_mkp,
                angkatan_pilihan=angkatan_pilihan,
            )
            current_praktikum = generate_praktikum(
                lab_kapasitas=lab_kapasitas,
                distribusi_praktikum=distribusi_prak,
                kelas_per_mkw=kelas_mkw,
            )
        except ScheduleError as e:
            error = str(e)

    return render_template(
        "index.html",
        schedule=current_schedule,
        praktikum=current_praktikum,
        error=error,
        prak_matrix=prak_matrix,
    )


# ---------- Export helpers ----------
def _style_workbook(wb):
    """Rapikan tiap sheet: header berwarna, border, freeze, lebar kolom."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    head_fill = PatternFill("solid", fgColor="4F46E5")
    head_font = Font(bold=True, color="FFFFFF")
    meta_fill = PatternFill("solid", fgColor="EEF2FF")
    thin = Side(style="thin", color="D5DAE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = center
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.border = border
                cell.alignment = center
                if i < 2:
                    cell.fill = meta_fill
                    cell.font = Font(bold=True)
        for col in ws.columns:
            letter = col[0].column_letter
            width = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 32)
        ws.freeze_panes = "C2"
        ws.row_dimensions[1].height = 22


def _build_pdf(path, schedule, praktikum):
    """Buat PDF landscape berisi tabel Kuliah & Praktikum."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    elems = []

    def add_table(title, rows):
        if not rows:
            return
        elems.append(Paragraph(f"<b>{title}</b>", styles["Heading2"]))
        elems.append(Spacer(1, 4))
        cols = list(rows[0].keys())
        data = [cols] + [[("" if r[c] == "-" else str(r[c])) for c in cols] for r in rows]
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5DAE5")),
            ("BACKGROUND", (0, 1), (1, -1), colors.HexColor("#EEF2FF")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (2, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        elems.append(tbl)
        elems.append(Spacer(1, 14))

    elems.append(Paragraph("<b>Jadwal Penjadwalan Kuliah Farmasi</b>", styles["Title"]))
    elems.append(Spacer(1, 8))
    add_table("Jadwal Kuliah", schedule)
    add_table("Jadwal Praktikum", praktikum)
    doc.build(elems)


@app.route("/export")
def export():
    global current_schedule, current_praktikum
    if not current_schedule:
        return "Belum ada jadwal untuk diekspor. Buat jadwal dulu.", 400
    file_path = os.path.join(tempfile.gettempdir(), "jadwal.xlsx")
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        pd.DataFrame(current_schedule).to_excel(writer, sheet_name="Kuliah", index=False)
        if current_praktikum:
            pd.DataFrame(current_praktikum).to_excel(writer, sheet_name="Praktikum", index=False)
        _style_workbook(writer.book)
    return send_file(file_path, as_attachment=True, download_name="jadwal.xlsx")


@app.route("/export/pdf")
def export_pdf():
    global current_schedule, current_praktikum
    if not current_schedule:
        return "Belum ada jadwal untuk diekspor. Buat jadwal dulu.", 400
    file_path = os.path.join(tempfile.gettempdir(), "jadwal.pdf")
    _build_pdf(file_path, current_schedule, current_praktikum)
    return send_file(file_path, as_attachment=True, download_name="jadwal.pdf")


if __name__ == "__main__":
    port = 5001
    if not getattr(sys, "frozen", False):
        # dev mode: hot reload
        app.run(debug=True, port=port)
    else:
        # desktop mode: native window (fallback ke browser bila gagal)
        url = f"http://127.0.0.1:{port}"
        threading.Thread(
            target=lambda: app.run(host="127.0.0.1", port=port, debug=False),
            daemon=True,
        ).start()
        try:
            import webview
            webview.create_window("Penjadwalan Kuliah Farmasi", url, width=1280, height=820)
            webview.start()
        except Exception:
            threading.Timer(1.0, lambda: webbrowser.open(url)).start()
            while True:
                threading.Event().wait(3600)
